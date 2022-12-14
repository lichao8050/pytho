# _*_ coding: utf-8 _*_
# @Time     : 2022/4/25 20:24
# @Author   : Mr_Li
# @FileName :
# Excel 读取为用例对象
from openpyxl import load_workbook

# excel = load_workbook("test_excel.xlsx")
# sheet = excel.worksheets[0]
# print(sheet.cell(2, 1).value)


class ExcelReader:
    '''不是针对普通Excel操作，专门针对测试用例表的  读、写'''

    def __init__(self, excel_path):
        # start_row = 5  从表的第5行开始读取
        self.start_row = 2
        self.excel_path = excel_path
        self.excel_file = load_workbook(excel_path)
        # sheet_1 = self.excel_file.worksheets[0]  从第一个sheet开始
        self.sheet_1 = self.excel_file.worksheets[0]

    def get_case_count(self):
        '''获取测试用例的数量
        ：return ：返回数量
        '''
        # 获取这张表总共有多少行
        rows_count = self.sheet_1.max_row
        # 有效用例数必须减去表头
        real_count = rows_count - (self.start_row - 1)
        return real_count

    def get_url(self, rows):
        '''获取接口URL， rows的参数是0开始，表示第1行'''
        # self.start_row + rows 起始行+想要获取的行， 3表示接口url所在的固定列是3
        return self.sheet_1.cell(self.start_row + rows, 3).value

    def get_method(self, rows):
        '''获取接口method， rows的参数是0开始，表示第1行'''
        # self.start_row + rows 起始行+想要获取的行， 3表示接口method所在的固定列是2
        return self.sheet_1.cell(self.start_row + rows, 2).value

    def get_headers(self, rows):
        '''获取接口headers， rows的参数是0开始，表示第1行'''
        # self.start_row + rows 起始行+想要获取的行， 4表示接口headers所在的固定列是4
        return self.sheet_1.cell(self.start_row + rows, 4).value

    def get_test_case_name(self, rows):
        '''获取接口用例名称， rows的参数是0开始，表示第1行'''
        # self.start_row + rows 起始行+想要获取的行， 1表示接口用例名称所在的固定列是1
        return self.sheet_1.cell(self.start_row + rows, 1).value

    def get_test_data(self, rows):
        '''获取接口请求参数， rows的参数是0开始，表示第1行'''
        # self.start_row + rows 起始行+想要获取的行， 5表示接口用例名称所在的固定列是5
        return self.sheet_1.cell(self.start_row + rows, 5).value

    def get_test_status_cood(self, rows):
        '''获取预期返回状态码， rows的参数是0开始，表示第1行'''
        # self.start_row + rows 起始行+想要获取的行， 6表示接口用例名称所在的固定列是6
        return self.sheet_1.cell(self.start_row + rows, 6).value

    def close_file(self,):
        '''关闭Excel'''
        self.excel_file.close()

    def set_pass_or_fail(self, rows, text):
        self.sheet_1.cell(self.start_row + rows, 6, value=text)